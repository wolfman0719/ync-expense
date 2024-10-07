import openpyxl
from openpyxl import Workbook
import iris
import pandas
import sys
import datetime

args = sys.argv

filename = args[1]

status = iris.cls('YNC.Expense')._KillExtent()
wb = openpyxl.load_workbook(filename)
ws = wb['input']
		
row_index = 0
for row in ws.iter_rows():
	row_index = row_index + 1
	if row_index < 5: continue
	month = ws.cell(row=row_index,column=2).value
	if (month == '' or month is None): continue
	day = ws.cell(row=row_index,column=3).value
	if (day == '' or day is None): continue
	paymentto = ws.cell(row=row_index,column=4).value
	accounts = ws.cell(row=row_index, column=5).value
	if accounts is None: accounts = '旅費交通費'
	amount = ws.cell(row=row_index, column=6).value
	if (amount == '' or amount is None): amount = 0
	description = ws.cell(row=row_index, column=7).value
	if description is None: description = 'no data'
	isjreimbursement = ws.cell(row=row_index,column=8).value
	onbehalf = ws.cell(row=row_index,column=9).value
	if isjreimbursement is None: isjreimbursement = ''
	if onbehalf is None: onbehalf = ''
	sql = iris.sql.prepare("insert into ync.expense(reportmonth,reportday,paymentto,accounts,amount,description,isjreimbursement,onbehalf) values(?,?,?,?,?,?,?,?)")
	sql.execute(month,day,paymentto,accounts,amount,description,isjreimbursement,onbehalf)

	sql = iris.sql.prepare("select description from ync.expenseitem where description = ?")
	rs = sql.execute(description)
	try:
		next(rs)
		exist = True
	except Exception:
		exist = False
	if exist is False:
		sql = iris.sql.prepare("insert into ync.expenseitem(paymentto,accounts,amount,description,onbehalf) values(?,?,?,?,?)")
		sql.execute(paymentto,accounts,amount,description,onbehalf)
	else:
		sql = iris.sql.prepare("update ync.expenseitem set paymentto=?,accounts=?,amount=?,onbehalf=? where description= ?")
		sql.execute(paymentto,accounts,amount,onbehalf,description)

wb.close()

wb = openpyxl.load_workbook(filename)
ws = wb['sorted']
		
itemline = iris.sql.exec("SELECT reportmonth, reportday, paymentto, accounts, amount, description, isjreimbursement, onbehalf FROM ync.expense order by reportmonth, reportday").dataframe()
	
linepos = 4
	  
for index,row in itemline.iterrows():
	rowline = list(row)
	paymentto  = rowline[2]
	accounts = rowline[3]
	amount = rowline[4]
	description = rowline[5]
	isjreimbursement = rowline[6]
	onbehalf = rowline[7]

	if (accounts is None or accounts == ''): accounts = '旅費交通費'

	sql = iris.sql.prepare("select paymentto, accounts, amount, onbehalf from ync.expenseitem where description = ?")
	rs = sql.execute(description)
	for index,row in enumerate(rs):
		paymentto = row[0]
		accounts = row[1]
		if (accounts is None or accounts == ''): accounts = '旅費交通費'
		if (amount == 0 or amount is None): amount = row[2]
		onbehalf = row[3]  
	
	linepos = linepos + 1
	ws.cell(row=linepos,column=2).value = rowline[0]
	ws.cell(row=linepos,column=3).value = rowline[1]
	ws.cell(row=linepos,column=4).value = paymentto
	ws.cell(row=linepos,column=5).value = accounts
	ws.cell(row=linepos,column=6).value = amount
	ws.cell(row=linepos,column=7).value = description
	ws.cell(row=linepos,column=8).value = isjreimbursement
	ws.cell(row=linepos,column=9).value = onbehalf
wb.save(filename)
wb.close()
