import openpyxl
from openpyxl import Workbook
import iris
import pandas
import sys
import datetime

args = sys.argv

filename = args[1]

year = datetime.datetime.today().year
reiwayear = year - 2018

status = iris.cls('YNC.Expense')._KillExtent()
wb = openpyxl.load_workbook(filename)
ws = wb['input']
		
row_index = 0
for row in ws.iter_rows():
	row_index = row_index + 1
	if row_index < 2: continue
	month = ws.cell(row=row_index,column=2).value
	if (month == '' or month is None): continue
	day = ws.cell(row=row_index,column=3).value
	if (day == '' or day is None): continue
	paymentto = ws.cell(row=row_index,column=4).value
	accounts = ws.cell(row=row_index, column=5).value
	if accounts is None: accounts = '旅費交通費'
	amount = ws.cell(row=row_index, column=6).value
	if (amount == '' or amount is None): amount = 0
	description = ws.cell(row=row_index, column=7).value
	sql = iris.sql.prepare("insert into ync.expense(reportmonth,reportday,paymentto,accounts,amount,description) values(?,?,?,?,?,?)")
	sql.execute(month,day,paymentto,accounts,amount,description)

wb.close()

wb = openpyxl.load_workbook(filename)
ws = wb['印刷用']
		
itemline = iris.sql.exec("select max(reportmd) from ync.expense").dataframe()

for index,row in itemline.iterrows():
	rowline = list(row)
	maxmd = rowline[0]
	maxmd = str(maxmd)
	if (len(maxmd) == 3): maxmd = '0' + maxmd
	maxmonth = maxmd[0:2]
	maxday = maxmd[2:4]

itemline = iris.sql.exec("select min(reportmd) from ync.expense").dataframe()

for index,row in itemline.iterrows():
	rowline = list(row)
	minmd = rowline[0]
	minmd = str(minmd)
	if (len(minmd) == 3): minmd = '0' + minmd
	minmonth = minmd[0:2]
	minday = minmd[2:4]

ws.cell(row=2,column=2).value = reiwayear
ws.cell(row=2,column=4).value = maxmonth
ws.cell(row=2,column=6).value = maxday

ws.cell(row=7,column=14).value = reiwayear
ws.cell(row=7,column=16).value = minmonth
ws.cell(row=7,column=18).value = minday
ws.cell(row=8,column=14).value = reiwayear
ws.cell(row=8,column=16).value = maxmonth
ws.cell(row=8,column=18).value = maxday

ws.cell(row=13,column=4).value = str(reiwayear) + '年' + minmonth + '月' + minday + '日'
ws.cell(row=13,column=12).value = str(reiwayear) + '年' + maxmonth + '月' + maxday + '日'

itemline = iris.sql.exec("select reportmonth, reportday, amount, description, paymentto from ync.expense order by reportmonth, reportday").dataframe()
	
linepos = 16

hotelamount = 0
hotelname = ''
nights = 0
	  
for index,row in itemline.iterrows():
	rowline = list(row)
	month = rowline[0]
	day = rowline[1]
	amount = rowline[2]
	description = rowline[3]
	paymentto = rowline[4]
	if description == 'HOTEL':
		hotelamount = hotelamount + amount
		hotelname = hotelname + paymentto + ' '
		nights = nights + 1 
		continue
	if paymentto == 'JAL':
		description = '飛行機　(' + description + ')'
	else:
		description = '電車　(' + description + ')'	
	linepos = linepos + 1
	if linepos > 26:
		print('# of lineitems exceeded the max limit') 
		continue
	ws.cell(row=linepos,column=1).value = str(month) + '/' + str(day)
	ws.cell(row=linepos,column=6).value = str(month) + '/' + str(day)
	ws.cell(row=linepos,column=10).value = description
	ws.cell(row=linepos,column=16).value = amount

ws.cell(row=30,column=8).value = hotelamount
ws.cell(row=30,column=2).value = hotelname
ws.cell(row=28,column=5).value = nights
wb.save(filename)
wb.close()
